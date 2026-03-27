"""
Unit tests for main.py — no real API calls, no credentials needed.
Run with: python -m pytest test_main.py -v
"""

import os
import json
import tempfile
from unittest.mock import MagicMock, patch, call
import pytest

# Provide dummy env vars before importing main
os.environ.setdefault("PAYPAL_CLIENT_ID", "test_client_id")
os.environ.setdefault("PAYPAL_CLIENT_SECRET", "test_client_secret")

import main  # noqa: E402  (import after env setup)


# ── _extract_row ──────────────────────────────────────────────────────────────

class TestExtractRow:
    def _item(self, **overrides):
        base = {
            "transaction_info": {
                "transaction_id": "TXN123",
                "transaction_initiation_date": "2024-01-15T10:00:00+0000",
                "transaction_amount": {"value": "50.00", "currency_code": "USD"},
                "transaction_status": "S",
            },
            "payer_info": {
                "email_address": "payer@example.com",
                "payer_name": {"given_name": "John", "surname": "Doe"},
            },
        }
        base.update(overrides)
        return base

    def test_full_row_extracted(self):
        row = main._extract_row(self._item())
        assert row["transaction_id"] == "TXN123"
        assert row["payer_email"] == "payer@example.com"
        assert row["payer_name"] == "John Doe"
        assert row["amount"] == "50.00"
        assert row["currency"] == "USD"
        assert row["status"] == "S"

    def test_no_payer_info_returns_none(self):
        item = self._item()
        item["payer_info"] = {}
        assert main._extract_row(item) is None

    def test_alternate_full_name_preferred(self):
        item = self._item()
        item["payer_info"]["payer_name"]["alternate_full_name"] = "Johnny D."
        row = main._extract_row(item)
        assert row["payer_name"] == "Johnny D."

    def test_missing_surname_graceful(self):
        item = self._item()
        item["payer_info"]["payer_name"] = {"given_name": "Jane"}
        row = main._extract_row(item)
        assert row["payer_name"] == "Jane"

    def test_empty_amount_handled(self):
        item = self._item()
        item["transaction_info"]["transaction_amount"] = {}
        row = main._extract_row(item)
        assert row["amount"] == ""
        assert row["currency"] == ""


# ── _fmt_date ─────────────────────────────────────────────────────────────────

class TestFmtDate:
    def test_format_output(self):
        from datetime import datetime
        dt = datetime(2024, 6, 15, 12, 30, 0)
        result = main._fmt_date(dt)
        assert result == "2024-06-15T12:30:00+0000"


# ── get_paypal_token ──────────────────────────────────────────────────────────

class TestGetPaypalToken:
    def test_returns_access_token(self):
        mock_resp = MagicMock()
        mock_resp.json.return_value = {"access_token": "abc123"}
        with patch("main.requests.post", return_value=mock_resp) as mock_post:
            token = main.get_paypal_token()
        assert token == "abc123"
        mock_post.assert_called_once()

    def test_uses_basic_auth_header(self):
        mock_resp = MagicMock()
        mock_resp.json.return_value = {"access_token": "tok"}
        with patch("main.requests.post", return_value=mock_resp) as mock_post:
            main.get_paypal_token()
        headers = mock_post.call_args[1]["headers"]
        assert headers["Authorization"].startswith("Basic ")

    def test_raises_on_http_error(self):
        mock_resp = MagicMock()
        mock_resp.raise_for_status.side_effect = Exception("401 Unauthorized")
        with patch("main.requests.post", return_value=mock_resp):
            with pytest.raises(Exception, match="401"):
                main.get_paypal_token()


# ── fetch_transactions ────────────────────────────────────────────────────────

class TestFetchTransactions:
    def _page(self, items, total_pages=1):
        mock = MagicMock()
        mock.json.return_value = {
            "transaction_details": items,
            "total_pages": total_pages,
        }
        return mock

    def _txn_item(self, tid="T1", email="a@b.com"):
        return {
            "transaction_info": {
                "transaction_id": tid,
                "transaction_initiation_date": "2026-01-20T00:00:00+0000",
                "transaction_amount": {"value": "10.00", "currency_code": "USD"},
                "transaction_status": "S",
            },
            "payer_info": {
                "email_address": email,
                "payer_name": {"given_name": "A", "surname": "B"},
            },
        }

    def test_single_page(self):
        item = self._txn_item()
        with patch("main.requests.get", return_value=self._page([item])):
            rows = main.fetch_transactions("token")
        assert len(rows) >= 1
        assert any(r["transaction_id"] == "T1" for r in rows)

    def test_pagination_fetches_all_pages(self):
        # Two pages within one window, then empty windows for remaining chunks
        page1 = self._page([self._txn_item("T1")], total_pages=2)
        page2 = self._page([self._txn_item("T2")], total_pages=2)
        empty = self._page([])
        with patch("main.requests.get", side_effect=[page1, page2, empty, empty, empty]):
            rows = main.fetch_transactions("token")
        assert {r["transaction_id"] for r in rows} == {"T1", "T2"}

    def test_empty_response(self):
        with patch("main.requests.get", return_value=self._page([])):
            rows = main.fetch_transactions("token")
        assert rows == []

    def test_items_without_payer_info_skipped(self):
        item = self._txn_item()
        item["payer_info"] = {}
        with patch("main.requests.get", return_value=self._page([item])):
            rows = main.fetch_transactions("token")
        assert rows == []

    def test_start_date_is_jan_15_2026(self):
        from datetime import datetime
        assert main.PAYPAL_START_DATE == datetime(2026, 1, 15)


# ── write_excel ───────────────────────────────────────────────────────────────

class TestWriteExcel:
    def _rows(self):
        return [
            {
                "transaction_id": "TXN1",
                "date": "2024-01-15",
                "payer_name": "Alice",
                "payer_email": "alice@example.com",
                "amount": "25.00",
                "currency": "USD",
                "status": "S",
            }
        ]

    def test_file_created(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            main.write_excel(self._rows(), path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0
        finally:
            os.unlink(path)

    def test_headers_written(self):
        from openpyxl import load_workbook
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            main.write_excel(self._rows(), path)
            wb = load_workbook(path)
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
            assert headers == ["Transaction ID", "Date", "Payer Name", "Payer Email", "Amount", "Currency", "Status"]
        finally:
            os.unlink(path)

    def test_data_row_written(self):
        from openpyxl import load_workbook
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            main.write_excel(self._rows(), path)
            wb = load_workbook(path)
            ws = wb.active
            assert ws.cell(row=2, column=1).value == "TXN1"
            assert ws.cell(row=2, column=4).value == "alice@example.com"
            assert ws.cell(row=2, column=5).value == 25.0  # stored as float
        finally:
            os.unlink(path)

    def test_invalid_amount_stored_as_string(self):
        from openpyxl import load_workbook
        rows = self._rows()
        rows[0]["amount"] = "N/A"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            main.write_excel(rows, path)
            wb = load_workbook(path)
            ws = wb.active
            assert ws.cell(row=2, column=5).value == "N/A"
        finally:
            os.unlink(path)

    def test_empty_rows_writes_headers_only(self):
        from openpyxl import load_workbook
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            main.write_excel([], path)
            wb = load_workbook(path)
            ws = wb.active
            assert ws.max_row == 1  # only header row
        finally:
            os.unlink(path)


# ── add_to_google_group ───────────────────────────────────────────────────────

class TestAddToGoogleGroup:
    def _service(self):
        svc = MagicMock()
        svc.members().insert().execute.return_value = {}
        return svc

    def test_successful_add_returns_true(self):
        svc = self._service()
        result = main.add_to_google_group(svc, "user@example.com")
        assert result is True

    def test_already_member_returns_false(self):
        svc = MagicMock()
        svc.members().insert().execute.side_effect = Exception("Member already exists")
        result = main.add_to_google_group(svc, "existing@example.com")
        assert result is False

    def test_409_conflict_returns_false(self):
        svc = MagicMock()
        svc.members().insert().execute.side_effect = Exception("409 conflict")
        result = main.add_to_google_group(svc, "existing@example.com")
        assert result is False

    def test_other_error_returns_false(self):
        svc = MagicMock()
        svc.members().insert().execute.side_effect = Exception("500 server error")
        result = main.add_to_google_group(svc, "user@example.com")
        assert result is False


# ── _build_google_service ─────────────────────────────────────────────────────

class TestBuildGoogleService:
    def test_missing_sa_file_returns_none(self):
        original = main.GOOGLE_SA_FILE
        main.GOOGLE_SA_FILE = "/nonexistent/path.json"
        result = main._build_google_service()
        main.GOOGLE_SA_FILE = original
        assert result is None

    def test_missing_admin_email_returns_none(self, tmp_path):
        sa_file = tmp_path / "sa.json"
        sa_file.write_text("{}")
        original_sa = main.GOOGLE_SA_FILE
        original_admin = main.GOOGLE_ADMIN_EMAIL
        main.GOOGLE_SA_FILE = str(sa_file)
        main.GOOGLE_ADMIN_EMAIL = ""
        result = main._build_google_service()
        main.GOOGLE_SA_FILE = original_sa
        main.GOOGLE_ADMIN_EMAIL = original_admin
        assert result is None
