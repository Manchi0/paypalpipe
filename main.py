"""
PayPal → Excel + Google Group
==============================
Fetches received PayPal payments, writes them to an Excel file,
and adds each payer to a Google Workspace admin group.

Setup:
  1. Copy .env.example to .env and fill in your credentials.
  2. pip install -r requirements.txt
  3. For Google: place your service_account.json in this directory
     and ensure domain-wide delegation is configured (see README section below).
  4. Run: python main.py

Google Admin SDK setup (one-time):
  - Google Cloud Console → IAM & Admin → Service Accounts → Create
  - Download JSON key, save as service_account.json
  - Enable domain-wide delegation on the service account
  - In admin.google.com → Security → API Controls → Domain-wide Delegation
    → Add Client ID with scope:
    https://www.googleapis.com/auth/admin.directory.group.member
  - Enable "Admin SDK API" in Google Cloud Console → APIs & Services → Library
"""

import os
import base64
import logging
from datetime import datetime, timedelta, timezone

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────

PAYPAL_CLIENT_ID     = os.environ["PAYPAL_CLIENT_ID"]
PAYPAL_CLIENT_SECRET = os.environ["PAYPAL_CLIENT_SECRET"]
PAYPAL_MODE          = os.getenv("PAYPAL_MODE", "sandbox")
PAYPAL_START_DATE    = datetime.strptime(
    os.getenv("PAYPAL_START_DATE", "2026-01-15"), "%Y-%m-%d"
)

PAYPAL_BASE = (
    "https://api-m.sandbox.paypal.com"
    if PAYPAL_MODE == "sandbox"
    else "https://api-m.paypal.com"
)

GOOGLE_SA_FILE    = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "service_account.json")
GOOGLE_ADMIN_EMAIL = os.environ.get("GOOGLE_ADMIN_EMAIL", "")
GOOGLE_GROUP_EMAIL = os.environ.get("GOOGLE_GROUP_EMAIL", "")

EXCEL_OUTPUT_PATH = os.getenv("EXCEL_OUTPUT_PATH", "transactions.xlsx")

# ── PayPal ────────────────────────────────────────────────────────────────────

def get_paypal_token() -> str:
    """Exchange client credentials for a PayPal Bearer token."""
    encoded = base64.b64encode(
        f"{PAYPAL_CLIENT_ID}:{PAYPAL_CLIENT_SECRET}".encode()
    ).decode()
    response = requests.post(
        f"{PAYPAL_BASE}/v1/oauth2/token",
        headers={
            "Authorization": f"Basic {encoded}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        data={"grant_type": "client_credentials"},
        timeout=15,
    )
    response.raise_for_status()
    token = response.json()["access_token"]
    log.info("PayPal token obtained.")
    return token


def _fmt_date(dt: datetime) -> str:
    """Format a datetime to PayPal's expected ISO 8601 format."""
    return dt.strftime("%Y-%m-%dT%H:%M:%S+0000")


def _fetch_window(token: str, start: datetime, end: datetime) -> list[dict]:
    """Fetch all paginated transactions within a single <=31-day window."""
    all_rows: list[dict] = []
    page = 1
    while True:
        params = {
            "start_date": _fmt_date(start),
            "end_date":   _fmt_date(end),
            "fields":     "transaction_info,payer_info",
            "page_size":  500,
            "page":       page,
            "transaction_status": "S",
        }
        response = requests.get(
            f"{PAYPAL_BASE}/v1/reporting/transactions",
            headers={"Authorization": f"Bearer {token}"},
            params=params,
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()

        for item in data.get("transaction_details", []):
            row = _extract_row(item)
            if row:
                all_rows.append(row)

        total_pages = data.get("total_pages", 1)
        log.info("Fetched page %d/%d (%d transactions so far)", page, total_pages, len(all_rows))

        if page >= total_pages:
            break
        page += 1
    return all_rows


def fetch_transactions(token: str) -> list[dict]:
    """
    Fetch all transactions from PAYPAL_START_DATE to now.
    Automatically chunks into <=31-day windows (PayPal API limit).
    """
    now   = datetime.now(timezone.utc).replace(tzinfo=None)
    start = PAYPAL_START_DATE

    all_rows: list[dict] = []
    window = timedelta(days=31)
    chunk_start = start

    while chunk_start < now:
        chunk_end = min(chunk_start + window, now)
        log.info("Fetching window %s → %s", _fmt_date(chunk_start), _fmt_date(chunk_end))
        all_rows.extend(_fetch_window(token, chunk_start, chunk_end))
        chunk_start = chunk_end

    log.info("Total transactions fetched: %d", len(all_rows))
    return all_rows


def _extract_row(item: dict) -> dict | None:
    """Pull the fields we care about from a transaction_details entry."""
    ti = item.get("transaction_info", {})
    pi = item.get("payer_info", {})

    # Skip if no payer info (e.g. internal transfers)
    if not pi:
        return None

    payer_name_obj = pi.get("payer_name", {})
    given  = payer_name_obj.get("given_name", "")
    surname = payer_name_obj.get("surname", "")
    full_name = payer_name_obj.get("alternate_full_name") or f"{given} {surname}".strip()

    amount_obj = ti.get("transaction_amount", {})

    return {
        "transaction_id": ti.get("transaction_id", ""),
        "date":           ti.get("transaction_initiation_date", ""),
        "payer_name":     full_name,
        "payer_email":    pi.get("email_address", ""),
        "amount":         amount_obj.get("value", ""),
        "currency":       amount_obj.get("currency_code", ""),
        "status":         ti.get("transaction_status", ""),
    }

# ── Excel ─────────────────────────────────────────────────────────────────────

HEADERS = ["Transaction ID", "Date", "Payer Name", "Payer Email", "Amount", "Currency", "Status"]
HEADER_BG  = "003087"   # PayPal navy
HEADER_FG  = "FFFFFF"

def write_excel(rows: list[dict], path: str) -> None:
    """Write transaction rows to a formatted .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_font = Font(bold=True, color=HEADER_FG)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    center      = Alignment(horizontal="center", vertical="center")

    # Header row
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # Data rows
    for row_num, txn in enumerate(rows, start=2):
        ws.cell(row=row_num, column=1, value=txn["transaction_id"])
        ws.cell(row=row_num, column=2, value=txn["date"])
        ws.cell(row=row_num, column=3, value=txn["payer_name"])
        ws.cell(row=row_num, column=4, value=txn["payer_email"])
        # Store amount as a number so Excel can sum/format it
        try:
            ws.cell(row=row_num, column=5, value=float(txn["amount"]))
        except (ValueError, TypeError):
            ws.cell(row=row_num, column=5, value=txn["amount"])
        ws.cell(row=row_num, column=6, value=txn["currency"])
        ws.cell(row=row_num, column=7, value=txn["status"])

    # Auto-width columns
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22

    ws.freeze_panes = "A2"  # Keep header visible when scrolling

    wb.save(path)
    log.info("Excel saved: %s", path)

# ── Google Admin ──────────────────────────────────────────────────────────────

def _build_google_service():
    """
    Build the Google Admin Directory service using a service account
    with domain-wide delegation, impersonating GOOGLE_ADMIN_EMAIL.
    Returns None if Google credentials are not configured.
    """
    if not os.path.exists(GOOGLE_SA_FILE):
        log.warning(
            "service_account.json not found — skipping Google Group step. "
            "Set GOOGLE_SERVICE_ACCOUNT_FILE and configure domain-wide delegation."
        )
        return None
    if not GOOGLE_ADMIN_EMAIL or not GOOGLE_GROUP_EMAIL:
        log.warning("GOOGLE_ADMIN_EMAIL or GOOGLE_GROUP_EMAIL not set — skipping Google Group step.")
        return None

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

        SCOPES = ["https://www.googleapis.com/auth/admin.directory.group.member"]
        creds = service_account.Credentials.from_service_account_file(
            GOOGLE_SA_FILE, scopes=SCOPES
        )
        delegated = creds.with_subject(GOOGLE_ADMIN_EMAIL)
        return build("admin", "directory_v1", credentials=delegated, cache_discovery=False)
    except ImportError:
        log.error("Google packages not installed. Run: pip install -r requirements.txt")
        return None


def add_to_google_group(service, email: str) -> bool:
    """
    Add a member to the configured Google Group.
    Returns True on success, False if already a member or on error.
    """
    try:
        service.members().insert(
            groupKey=GOOGLE_GROUP_EMAIL,
            body={"email": email, "role": "MEMBER", "delivery_settings": "ALL_MAIL"},
        ).execute()
        log.info("Added to Google Group: %s", email)
        return True
    except Exception as exc:
        err_str = str(exc)
        if "Member already exists" in err_str or "409" in err_str:
            log.info("Already in group (skipped): %s", email)
        else:
            log.warning("Could not add %s: %s", email, exc)
        return False

# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    log.info("=== PayPal Flow starting (mode: %s) ===", PAYPAL_MODE)

    # 1. Fetch transactions from PayPal
    token = get_paypal_token()
    rows  = fetch_transactions(token)

    if not rows:
        log.info("No transactions found in the last %d days.", PAYPAL_DAYS_BACK)
        return

    # 2. Write to Excel
    write_excel(rows, EXCEL_OUTPUT_PATH)

    # 3. Add payers to Google Group
    service = _build_google_service()
    if service:
        seen_emails: set[str] = set()
        added = skipped = 0
        for txn in rows:
            email = txn["payer_email"].strip().lower()
            if not email or email in seen_emails:
                continue
            seen_emails.add(email)
            if add_to_google_group(service, email):
                added += 1
            else:
                skipped += 1
        log.info("Google Group: %d added, %d skipped", added, skipped)
    else:
        log.info("Google Group step skipped (not configured).")

    log.info("=== Done. ===")
    log.info("  Excel:  %s  (%d rows)", EXCEL_OUTPUT_PATH, len(rows))
    if GOOGLE_GROUP_EMAIL:
        log.info("  Group:  %s", GOOGLE_GROUP_EMAIL)


if __name__ == "__main__":
    main()
